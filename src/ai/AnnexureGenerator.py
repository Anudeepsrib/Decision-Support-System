"""
AnnexureGenerator.py
One-Click Petition Generator: Produces KSERC-standard Annexure Tables
in Excel (.xlsx) format, perfectly synced with the AI's computations.
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class AnnexureGenerator:
    """
    Generates KSERC-standard Annexure Tables in Excel format from
    the Phase 1 Rule Engine's audit trace outputs.
    """

    HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF') if HAS_OPENPYXL else None
    HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') if HAS_OPENPYXL else None
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    ) if HAS_OPENPYXL else None

    def __init__(self, output_dir: str = 'output'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_annexure(self, audit_results: list[dict], petition_title: str = "ARR Truing-Up") -> str:
        """
        Generates a KSERC-standard Excel Annexure workbook from audit trace data.

        Args:
            audit_results: List of AuditObject dicts from the Phase 1 engine.
            petition_title: Title for the petition header.

        Returns:
            Path to the generated .xlsx file.
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Sheet 1: Variance Summary (Annexure-I)
        self._create_variance_summary(wb, audit_results, petition_title)

        # Sheet 2: Gain/Loss Sharing (Annexure-II)
        self._create_sharing_detail(wb, audit_results)

        # Sheet 3: Anomaly Flags (Annexure-III)
        self._create_anomaly_sheet(wb, audit_results)

        # Remove default empty sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        filename = f"KSERC_Annexure_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath

    def _create_variance_summary(self, wb, results, title):
        ws = wb.create_sheet("Annexure-I Variance Summary")

        # Title Row
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{title} — Variance Summary Statement"
        ws['A1'].font = Font(name='Calibri', bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:G2')
        ws['A2'] = f"As per Order dated 30.06.2025 | Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws['A2'].font = Font(name='Calibri', italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['S.No', 'Cost Head', 'Category', 'Approved (Rs.)', 'Actual (Rs.)', 'Variance (Rs.)', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER

        # Data Rows
        for i, result in enumerate(results, 1):
            row = i + 4
            variance = result.get('variance_amount', 0)
            status = 'Gain' if variance >= 0 else 'Disallowed' if result.get('disallowed_variance', 0) > 0 else 'Pass-Through'

            data = [
                i,
                result.get('cost_head', ''),
                result.get('variance_category', ''),
                result.get('approved_amount', 0),
                result.get('actual_amount', 0),
                variance,
                status
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal='center')
                # Highlight losses in red
                if col == 7 and status == 'Disallowed':
                    cell.font = Font(color='FF0000', bold=True)

        # Auto-size columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_sharing_detail(self, wb, results):
        ws = wb.create_sheet("Annexure-II Sharing Detail")

        ws.merge_cells('A1:F1')
        ws['A1'] = "Gain/Loss Sharing Computation — Regulation 9.2"
        ws['A1'].font = Font(name='Calibri', bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['S.No', 'Cost Head', 'Variance (Rs.)', 'Disallowed (Rs.)', 'Passed Through (Rs.)', 'Logic Applied']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER

        for i, result in enumerate(results, 1):
            row = i + 3
            data = [
                i,
                result.get('cost_head', ''),
                result.get('variance_amount', 0),
                result.get('disallowed_variance', 0),
                result.get('passed_through_variance', 0),
                result.get('logic_applied', '')
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.BORDER

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def _create_anomaly_sheet(self, wb, results):
        ws = wb.create_sheet("Annexure-III AI Flags")

        ws.merge_cells('A1:D1')
        ws['A1'] = "AI Prudence Check — Anomaly Flags"
        ws['A1'].font = Font(name='Calibri', bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['S.No', 'Cost Head', 'Engine Version', 'Flags']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER

        for i, result in enumerate(results, 1):
            row = i + 3
            metadata = result.get('metadata', {})
            flags = ', '.join(metadata.get('flags', [])) or 'None'
            data = [
                i,
                result.get('cost_head', ''),
                metadata.get('engine_version', 'Unknown'),
                flags
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.BORDER
                if col == 4 and flags != 'None':
                    cell.font = Font(color='FF0000', bold=True)

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 24


# --- Demo Execution ---
if __name__ == "__main__":
    print("Generating KSERC Annexure Tables...\n")

    # Simulated Phase 1 audit results
    sample_audit_results = [
        {
            "cost_head": "O&M",
            "variance_category": "Controllable",
            "approved_amount": 180000000,
            "actual_amount": 150000000,
            "variance_amount": 30000000,
            "disallowed_variance": 0,
            "passed_through_variance": 10000000,
            "logic_applied": "Controllable Gain (Savings) split 2/3 Utility, 1/3 Consumer.",
            "metadata": {"engine_version": "Phase1-v1.0.0", "flags": []}
        },
        {
            "cost_head": "O&M",
            "variance_category": "Controllable",
            "approved_amount": 150000000,
            "actual_amount": 180000000,
            "variance_amount": -30000000,
            "disallowed_variance": 30000000,
            "passed_through_variance": 0,
            "logic_applied": "Controllable Loss fully borne by Utility (Disallowed).",
            "metadata": {"engine_version": "Phase1-v1.0.0", "flags": ["HIGH_ANOMALY_FLAG"]}
        },
        {
            "cost_head": "Power_Purchase",
            "variance_category": "Uncontrollable",
            "approved_amount": 400000000,
            "actual_amount": 450000000,
            "variance_amount": -50000000,
            "disallowed_variance": 0,
            "passed_through_variance": -50000000,
            "logic_applied": "Uncontrollable Variance fully passed through to Consumer.",
            "metadata": {"engine_version": "Phase1-v1.0.0", "flags": []}
        }
    ]

    generator = AnnexureGenerator(output_dir='output')
    filepath = generator.generate_annexure(sample_audit_results, "FY 2024-25 ARR Truing-Up")
    print(f"✅ KSERC Annexure generated: {filepath}")
